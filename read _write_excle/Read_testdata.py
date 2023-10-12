from openpyxl import *
wb=load_workbook(filename="employee.xlsx")
ws=wb.active


row=ws.max_row
col=ws.max_column
print(row)
print(col)
for r in range(1,row+1):
    for c in range(1,col+1):
        print(ws.cell(r,c).value,end=" ")

    print('')
