from openpyxl import *
wb=load_workbook(filename="emp.xlsx")
ws=wb.active

print(ws["A1"].value,end=" ")
print(ws["B1"].value)
print(ws["A2"].value,end=" ")
print(ws["B2"].value)

row=ws.max_row
col=ws.max_column
print(row)
print(col)
for r in range(1,row+1):
    for c in range(1,col+1):
        print(ws.cell(r,c).value,end=" ")

    print('\n')
