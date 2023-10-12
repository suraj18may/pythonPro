from openpyxl import *
wb=load_workbook("emp.xlsx")
ws=wb.active
ws["A1"]="Name"
ws["B1"]="EMPId"
ws["C1"]="salary"
ws["A2"].value="Mrinal"
ws["B2"].value="101"
ws["C2"]="70000"

wb.save("emp.xlsx")